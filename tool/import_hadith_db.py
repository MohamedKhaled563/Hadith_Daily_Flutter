"""Convert `assets/data/Hadith App DB.xlsx` into the JSON the app bundles.

Run from the project root after the workbook changes:

    python tool/import_hadith_db.py

Writes:
    assets/data/hadiths.json   — 42 hadiths (matn, explanation, narrator bio)
    assets/data/insights.json  — the daily messages

This is a transcription, not a rewrite: every Arabic string is copied verbatim
from the workbook. The only derived field is the title for hadiths 23-42, which
the workbook does not supply — see `derive_title`.

Requires: pip install openpyxl
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "assets" / "data" / "Hadith App DB.xlsx"
HADITHS_OUT = ROOT / "assets" / "data" / "hadiths.json"
INSIGHTS_OUT = ROOT / "assets" / "data" / "insights.json"

SHEET_MESSAGES = "رسائل"
SHEET_DATABASE = "قاعدة البيانات"

# The workbook labels message categories in English; the app is Arabic-only.
CATEGORY_AR = {
    "Short Tip": "لمسة قصيرة",
    "Practical Advice": "إرشاد عملي",
    "Reflection": "تأمّل",
    "Spiritual Reminder": "تذكير إيماني",
    "Gentle Warning": "تنبيه لطيف",
}

DEFAULT_REFERENCE = "من الأربعين النووية"

# Harakat, tatweel and other combining marks. The workbook mixes fully
# vocalised text ("رَوَاهُ مُسْلِمٌ") with bare text ("رواه مسلم"), so anything
# that matches or measures Arabic has to look past these.
DIACRITICS = "ؐ-ًؚ-ٰٟۖ-ۭـ"
DIACRITIC_RE = re.compile(f"[{DIACRITICS}]")


def undiacritised(text: str) -> str:
    return DIACRITIC_RE.sub("", text)


def flexible(word: str) -> str:
    """Regex matching `word` whether or not it carries diacritics."""
    return f"[{DIACRITICS}]*".join(re.escape(ch) for ch in word)


_ATTRIBUTION = "|".join(
    flexible(w)
    for w in ("رواه", "أخرجه", "رواة", "متفق", "حديث حسن", "حديث صحيح")
)

# Attribution or grading clause closing the matn: "رواه البخاري ومسلم",
# "رَوَاهُ التِّرْمِذِيُّ", "حديثٌ حسنٌ، رواه الترمذي وغيره".
TRAILING_REFERENCE_RE = re.compile(
    f"(?:{_ATTRIBUTION})[^\\n\"«»]{{0,60}}$",
    re.MULTILINE,
)

# Same clause anywhere in the matn, stopping at the first punctuation. Used
# when the attribution sits mid-text — hadith 5 attributes and then adds a
# variant narration, so the attribution never reaches the end.
INLINE_REFERENCE_RE = re.compile(
    f"(?:{_ATTRIBUTION})[^\\n،.؛\"«»]{{0,60}}"
)


# Particles and conjunctions that should not be the last word of a title.
CONNECTORS = {
    "و", "أو", "ثم", "ف", "على", "عن", "من", "في", "إلى", "حتى", "مع",
    "أن", "إن", "أنه", "إنه", "ما", "لا", "كل", "وكل", "بكل", "لكل",
    "الذي", "التي", "هو", "هي", "قد", "لقد", "يا", "ب", "ل", "ك",
    "وهو", "وهي", "وما", "ولا", "فإن", "وإن", "له", "لها", "به", "بها",
    "ذلك", "هذا", "هذه",
}


def clean(value) -> str:
    """Normalise whitespace without touching the Arabic itself."""
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def sections(markdown: str) -> list[tuple[str, str]]:
    """Split the elaboration markdown into (heading, body) pairs, in order."""
    parts = re.split(r"^#{2,}\s*(.+?)\s*$", markdown, flags=re.MULTILINE)
    # parts[0] is any preamble before the first heading.
    return [
        (clean(parts[i]), clean(parts[i + 1]))
        for i in range(1, len(parts) - 1, 2)
    ]


def join_sections(pairs: list[tuple[str, str]], prefix: str) -> str:
    """Concatenate every section whose heading starts with `prefix`.

    Several hadiths split their matn or commentary across
    "... - الجزء الأول / الثاني / الثالث" parts; those rejoin in sheet order.
    """
    bodies = [body for heading, body in pairs if heading.startswith(prefix)]
    return "\n\n".join(b for b in bodies if b).strip()


def extract_reference(matn: str) -> tuple[str, str]:
    """Pull the attribution out of the matn.

    Returns (matn, reference). The matn is only trimmed when the attribution
    actually closes it — hadith 5 attributes mid-text and then adds a variant
    narration ("وفي رواية لمسلم ..."), which has to stay in the text.
    """
    match = TRAILING_REFERENCE_RE.search(matn)
    if match:
        reference = clean(match.group(0)).strip(" .,،\"«»")
        return clean(matn[: match.start()]), reference or DEFAULT_REFERENCE

    # Attribution sits mid-text: take it as the reference but leave the matn
    # intact, since what follows it is still part of the narration.
    inline = list(INLINE_REFERENCE_RE.finditer(matn))
    if inline:
        reference = clean(inline[0].group(0)).strip(" .,،\"«»")
        return matn, reference or DEFAULT_REFERENCE

    return matn, DEFAULT_REFERENCE


def derive_title(matn: str, arabic_text: str) -> str:
    """Title for the hadiths the workbook does not name (23-42).

    The workbook's own titles for 1-22 are the hadith's opening words
    ("لا تغضب", "دع ما يريبك إلى ما لا يريبك"), which is the conventional way
    to refer to a hadith. This follows that convention rather than inventing a
    theme: it quotes the opening clause of the text and stops at a natural
    break.
    """
    source = clean(arabic_text) or clean(matn)

    # Drop any narration chain so the title starts at the Prophetic words.
    source = re.sub(r'^.*?(?:قَالَ|قال)\s*:?\s*"?', "", source, count=1).strip()
    source = source.lstrip('"«: ').strip()

    # Measure on the undiacritised form: a vocalised phrase carries roughly
    # twice the code points for the same visual width, so counting raw
    # characters truncates vocalised titles mid-word.
    def width(text: str) -> int:
        return len(undiacritised(text))

    # Prefer stopping at the first clause boundary — that yields a whole
    # phrase, which is what the workbook's own titles are.
    clause = re.split(r"[،.:؛]|\s\"|\s«", source, maxsplit=1)[0].strip()

    if 8 <= width(clause) <= 45:
        return clause.strip(' "«»')

    words = source.split()
    title = ""
    for word in words:
        candidate = f"{title} {word}".strip()
        if width(candidate) > 42:
            break
        title = candidate

    if not title:
        return (clause or "حديث نبوي شريف").strip(' "«»')

    # A hard cut can leave a dangling conjunction or preposition
    # ("... وَكُلِّ"). Drop those so the title ends on a content word.
    kept = title.split()
    while len(kept) > 3 and undiacritised(kept[-1]).strip() in CONNECTORS:
        kept.pop()
    title = " ".join(kept)

    # The workbook itself marks an abridged title this way (see hadith 22).
    truncated = width(title) < width(" ".join(words))
    return title.strip(' "«»') + ("…" if truncated else "")


def main() -> int:
    if not WORKBOOK.exists():
        print(f"error: workbook not found at {WORKBOOK}", file=sys.stderr)
        return 1

    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)

    # ---- messages sheet: titles + daily insights ---------------------------
    messages_sheet = workbook[SHEET_MESSAGES]
    message_rows = [
        row
        for row in messages_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None and clean(row[3])
    ]

    titles: dict[int, str] = {}
    insights: list[dict] = []

    for number, title, category, arabic, english, themes, keywords in message_rows:
        num = int(number)
        if clean(title):
            titles.setdefault(num, clean(title))

        insights.append(
            {
                "hadithNumber": num,
                "arabic": clean(arabic),
                "english": clean(english),
                "category": CATEGORY_AR.get(clean(category), clean(category)),
                "themes": clean(themes),
                "keywords": clean(keywords),
            }
        )

    # ---- database sheet: the hadiths themselves ---------------------------
    database_sheet = workbook[SHEET_DATABASE]
    hadiths: list[dict] = []
    derived_titles: list[int] = []
    missing_reference: list[int] = []

    for row in database_sheet.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue

        number = int(row[1])
        arabic_text = clean(row[2])
        pairs = sections(clean(row[3]))

        matn = join_sections(pairs, "نص الحديث الكامل") or arabic_text
        explanation = join_sections(pairs, "الشرح")
        matn, reference = extract_reference(matn)

        if reference == DEFAULT_REFERENCE:
            missing_reference.append(number)

        title = titles.get(number)
        if not title:
            title = derive_title(matn, arabic_text)
            derived_titles.append(number)

        hadiths.append(
            {
                "number": number,
                "title": title,
                "text": matn,
                "reference": reference,
                "explanation": explanation,
                # The workbook carries no per-hadith lesson list. Left empty
                # rather than invented; the detail screen hides the card.
                "keyLessons": [],
                "narratorBio": clean(row[4]),
            }
        )

    hadiths.sort(key=lambda h: h["number"])
    insights.sort(key=lambda i: (i["hadithNumber"], i["arabic"]))

    for path, payload in ((HADITHS_OUT, hadiths), (INSIGHTS_OUT, insights)):
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    covered = sorted({i["hadithNumber"] for i in insights})
    print(f"hadiths  : {len(hadiths):>3}  -> {HADITHS_OUT.relative_to(ROOT)}")
    print(f"insights : {len(insights):>3}  -> {INSIGHTS_OUT.relative_to(ROOT)}")
    print()
    print(f"titles taken from workbook : {len(hadiths) - len(derived_titles)}")
    print(f"titles derived from matn   : {len(derived_titles)} {derived_titles}")
    print(f"default reference used for : {missing_reference or 'none'}")
    print(f"insight coverage           : hadiths {covered[0]}-{covered[-1]} "
          f"({len(covered)} of {len(hadiths)})")

    uncovered = [h["number"] for h in hadiths if h["number"] not in set(covered)]
    if uncovered:
        print(f"NO daily messages for      : {uncovered}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
